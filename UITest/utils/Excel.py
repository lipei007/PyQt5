from openpyxl import load_workbook


# 从文件读取数据
def read(path):
    xls = load_workbook(filename=path)
    # print(xls.sheetnames)
    sheet = xls["Sheet1"]

    # 处理列合并
    merges = sheet.merged_cells.ranges
    merge_rows = []
    for rg in merges:
        r = rg.start_cell.row
        merge_rows.append(r)

    # 行遍历
    data_rows = []
    row = 0
    for r in sheet.iter_rows():
        row += 1
        if merge_rows.__contains__(row):
            continue
        data_rows.append(r)
    return data_rows


# 完全将Excel数据转成字符串二维数组
def convert_excel_data(rows):
    arr = []
    for row in rows:
        tmp_row = []
        for cell in row:
            value = cell.value
            if value is None:
                value = ''
            value = f'{value}'
            value = value.strip()
            tmp_row.append(value)
        arr.append(tmp_row)
    return arr


# 过滤不符合条件的行
def filter_rows(rows, rm_index_array):
    res = []
    le = len(rows)
    for i in range(0, le):
        if rm_index_array.__contains__(i):
            continue
        res.append(rows[i])
    return res
