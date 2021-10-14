from openpyxl import load_workbook
from database.Database import DataBasePool


# 从文件读取数据
def read(path):
    xls = load_workbook(filename=path)
    # print(xls.sheetnames)
    sheet = xls["Sheet1"]

    # 处理列合并
    merges = sheet.merged_cells.ranges
    merge_rows = []
    for rg in merges:
        r = rg.start_cell.row
        merge_rows.append(r)

    # 行遍历
    data_rows = []
    row = 0
    for r in sheet.iter_rows():
        row += 1
        if merge_rows.__contains__(row):
            continue
        data_rows.append(r)
    return data_rows


# 过滤不符合条件的行
def filter_rows(rows, rm_index_array):
    res = []
    le = len(rows)
    for i in range(0, le):
        if rm_index_array.__contains__(i):
            continue
        res.append(rows[i])
    return res


# 数据库连接池
mysql_pool = DataBasePool(host='localhost', port=3306, user='root', password='abc123456', db='load_emu')


# 查询表的所有字段
def query_all_columns(table):
    return mysql_pool.query(f'select COLUMN_NAME from information_schema.COLUMNS where table_name = \'{table}\';')


# 将数据插入数据库
def insert_rows_to_mysql(rows, maps):
    for r in rows:
        insert_row_to_mysql(r=r, maps=maps)


# 执行插入
def insert_row_to_mysql(r, maps):
    cols = len(r)

    keys = []
    values = []

    for j in range(0, cols):
        cell = r[j]
        value = cell.value
        key = maps.get(j)
        if key is not None:
            keys.append(key)

            if value is None:
                value = ""
            value = value.strip()
            values.append(f'\'{value}\'')

    if len(keys) > 0:
        sql = f'insert into us_data ({",".join(keys)}) values ({",".join(values)});'
        mysql_pool.execute(sql=sql)
